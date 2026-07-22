"""Generador de memoria de cálculo Excel con identidad DML.

Uso:
    from templates.plantilla_dml import nuevo_libro
    wb = nuevo_libro("P25XX-PR-MC-001", "Caída de presión línea L-101", "CLIENTE", "0")
    ws = wb.active
    ws["A10"] = "Parámetro"
    wb.save("memoria.xlsx")
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

VERDE_DML = "2E7D32"
GRIS_HDR = "E8E8E8"

_thin = Side(style="thin", color="808080")
BORDE = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def nuevo_libro(codigo: str, titulo: str, cliente: str, revision: str) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Memoria"

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 35

    ws.merge_cells("A1:E3")
    c = ws["A1"]
    c.value = f"DML INGENIERÍA\n{codigo}\n{titulo}"
    c.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=VERDE_DML)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws["A5"] = "Cliente:"
    ws["B5"] = cliente
    ws["D5"] = "Revisión:"
    ws["E5"] = revision
    for cell in ("A5", "D5"):
        ws[cell].font = Font(bold=True)

    headers = ["Ítem", "Descripción", "Valor", "Unidad", "Fuente / Referencia"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=8, column=i, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=GRIS_HDR)
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDE

    return wb
