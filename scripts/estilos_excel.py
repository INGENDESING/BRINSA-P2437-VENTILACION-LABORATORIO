#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
estilos_excel.py — Módulo ÚNICO de formato corporativo DML para los libros
Excel generados del proyecto (CAL-001, DTS-001/002/003, LIS-001).

Reglas vigentes (instrucción del cliente, 2026-07-28):
  - Hoja A3, orientación horizontal, ajuste a una página de ancho.
  - Times New Roman 28 en TODO el contenido generado (fila 9 en adelante);
    jerarquía por negrita y color, no por tamaño.
  - Encabezados de tabla en verde claro corporativo DML (sin azul).
  - Exactamente UNA fila en blanco entre bloques (gestionado por los
    generadores, no por este módulo).
  - Tablas distribuidas en el ancho total del encabezado corporativo (A:O).

El encabezado corporativo (filas 1-7) y la PORTADA provienen de las
plantillas `FormatosDocumentos/*.xlsx` y NO se modifican aquí.

Uso:  from estilos_excel import *  (los generadores agregan scripts/ a sys.path)
"""

import math

from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.properties import PageSetupProperties

# ===================== PALETA CORPORATIVA DML =====================
VERDE_CLARO = "C6E0B4"          # relleno de encabezados de tabla
VERDE_OSCURO = "375623"         # texto sobre verde claro y títulos de sección
AMARILLO_ENTRADA = "FFF2CC"     # celdas de entrada modificables (CAL)
VERDE_RESULTADO = "C6EFCE"      # celdas de resultado (CAL)
VERDE_RESULTADO_FUENTE = "006100"
GRIS_FORMULA = "E7E6E6"         # celdas de fórmula (CAL)
GRIS_NOTA = "404040"            # notas numeradas (DTS)

# ===================== TIPOGRAFÍA: TIMES NEW ROMAN 28 =====================
TAMANO = 28
FUENTE = "Times New Roman"

header_fill = PatternFill(start_color=VERDE_CLARO, end_color=VERDE_CLARO, fill_type="solid")
header_font = Font(name=FUENTE, size=TAMANO, bold=True, color=VERDE_OSCURO)
title_font = Font(name=FUENTE, size=TAMANO, bold=True, color=VERDE_OSCURO)      # títulos #
section_font = Font(name=FUENTE, size=TAMANO, bold=True, color=VERDE_OSCURO)    # títulos ##
subtitle_font = Font(name=FUENTE, size=TAMANO, bold=True, color=VERDE_OSCURO)   # títulos ###
caption_font = Font(name=FUENTE, size=TAMANO, bold=True)                        # leyendas Tabla N.
nota_font = Font(name=FUENTE, size=TAMANO, italic=True, color=GRIS_NOTA)        # notas numeradas
body_font = Font(name=FUENTE, size=TAMANO)
input_fill = PatternFill(start_color=AMARILLO_ENTRADA, end_color=AMARILLO_ENTRADA, fill_type="solid")
input_font = Font(name=FUENTE, size=TAMANO, bold=True)
result_fill = PatternFill(start_color=VERDE_RESULTADO, end_color=VERDE_RESULTADO, fill_type="solid")
result_font = Font(name=FUENTE, size=TAMANO, bold=True, color=VERDE_RESULTADO_FUENTE)
formula_fill = PatternFill(start_color=GRIS_FORMULA, end_color=GRIS_FORMULA, fill_type="solid")
formula_font = Font(name=FUENTE, size=TAMANO, bold=True, color=VERDE_OSCURO)

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Alineaciones unificadas por rol
AL_TEXTO = Alignment(horizontal='left', vertical='top', wrap_text=True)
AL_CENTRO = Alignment(horizontal='center', vertical='center', wrap_text=True)
AL_HEADER = Alignment(horizontal='center', vertical='center', wrap_text=True)
AL_PARRAFO = Alignment(horizontal='left', vertical='top', wrap_text=True)

# ===================== GEOMETRÍA (A3 horizontal, TNR 28) =====================
N_COLS = 15                     # layout unificado: todas las tablas abarcan A:O
CHAR_POR_UNIDAD_ANCHO = 0.9     # caracteres TNR 28 por unidad de ancho Excel
ALTURA_LINEA = 50               # altura de fila por línea de texto (pt)
ALTURA_MINIMA = 50              # altura mínima de cualquier fila de contenido

# Ancho de columna uniforme (instrucción del cliente, 2026-07-28): 35 en las
# 15 columnas A:O de las hojas de contenido. Las portadas conservan la plantilla.
ANCHO_COLUMNAS = [35] * N_COLS


def aplicar_ancho_columnas(ws):
    """Fija anchos proporcionales para las 15 columnas del documento."""
    for i, w in enumerate(ANCHO_COLUMNAS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def spans_tabla(n_cols):
    """Distribuye n_cols columnas lógicas en N_COLS columnas físicas (A:O).

    base = N_COLS // n_cols; el residuo se reparte en las primeras columnas.
    """
    if n_cols > N_COLS:
        raise ValueError(f"Tabla con {n_cols} columnas supera el layout A:{get_column_letter(N_COLS)}")
    base = N_COLS // n_cols
    extra = N_COLS % n_cols
    return [base + (1 if i < extra else 0) for i in range(n_cols)]


def forzar_times_new_roman(ws, min_row=1, max_row=None, min_col=1, max_col=N_COLS):
    """Sobrescribe el nombre de fuente conservando tamaño, negrita, cursiva, color y subrayado."""
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            f = cell.font
            cell.font = Font(
                name=FUENTE,
                size=f.size,
                bold=f.bold,
                italic=f.italic,
                underline=f.underline,
                color=f.color,
            )


def ajustar_alturas_filas(ws, primera, ultima):
    """Ajusta la altura de cada fila en función del texto más largo de la fila,
    considerando el ancho real (incluyendo celdas combinadas) y letra 28 pt."""
    for r in range(primera, ultima + 1):
        max_lineas = 1
        for c in range(1, N_COLS + 1):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            val = cell.value
            if not val:
                continue
            # Buscar rango combinado al que pertenezca la celda
            rng = None
            for m in ws.merged_cells.ranges:
                if cell.coordinate in m:
                    rng = m
                    break
            if rng:
                _, sc, _, ec = range_boundaries(str(rng))
                ancho = sum(ws.column_dimensions[get_column_letter(col)].width or 8.43
                            for col in range(sc, ec + 1))
                _, sr, _, _ = range_boundaries(str(rng))
                filas_combinadas = r - sr + 1
            else:
                ancho = ws.column_dimensions[get_column_letter(c)].width or 8.43
                filas_combinadas = 1
            chars_linea = max(1.0, ancho / CHAR_POR_UNIDAD_ANCHO)
            lineas_totales = math.ceil(len(str(val)) / chars_linea)
            lineas_por_fila = math.ceil(lineas_totales / filas_combinadas)
            max_lineas = max(max_lineas, lineas_por_fila)
        ws.row_dimensions[r].height = max(ALTURA_MINIMA, max_lineas * ALTURA_LINEA)


def configurar_pagina_a3(ws):
    """Configura la hoja en A3 horizontal, ajustada a una página de ancho."""
    ws.page_setup.paperSize = 8  # A3
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    else:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
