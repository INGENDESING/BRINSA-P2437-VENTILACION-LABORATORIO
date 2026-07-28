#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
generar_lis.py — Convierte el listado de equipos Markdown
(`Investigacion/Sistemas/listado_equipos.md`) a un libro Excel con la plantilla
corporativa `FormatosDocumentos/LIS.xlsx`.

El libro resultante tiene EXACTAMENTE 2 hojas:
  - PORTADA: portada corporativa diligenciada (BO1 = código, Z5 = título).
  - LISTA: encabezado corporativo (filas 1-7, logos y fórmulas =PORTADA!...),
    título del documento en fila 8 y el contenido del .md apilado desde la
    fila 9, con la tabla BOQ distribuida en el ancho total del encabezado
    (A:O, 15 columnas) para evitar huecos (encabezado verde claro DML, bordes
    thin, alineación por rol).

Salida: build/lis/P2437-HV-LIS-001 REV0.xlsx (intermedio que scripts/emitir.py
copia a Emisiones/4.0 HV-LISTADOS/).

Uso:  python scripts/generar_lis.py
"""

import math
import re
import sys
from copy import copy, deepcopy
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

# Módulo único de formato corporativo DML (A3 horizontal, Times New Roman 28,
# encabezados verde claro, sin azul). Ver scripts/estilos_excel.py.
sys.path.insert(0, str(Path(__file__).resolve().parent))
from estilos_excel import (  # noqa: E402
    N_COLS, AL_TEXTO, AL_HEADER, AL_PARRAFO,
    header_fill, header_font, title_font, section_font, subtitle_font,
    caption_font, body_font, thin_border,
    CHAR_POR_UNIDAD_ANCHO, ANCHO_COLUMNAS, ALTURA_MINIMA,
    aplicar_ancho_columnas, forzar_times_new_roman,
    ajustar_alturas_filas, configurar_pagina_a3,
)

ROOT = Path(__file__).resolve().parent.parent
PLANTILLA = ROOT / "FormatosDocumentos" / "LIS.xlsx"
SALIDA = ROOT / "build" / "lis"
FILA_INICIO = 9  # contenido tras encabezado corporativo (1-7), título (8)

# (archivo .md fuente, código del documento, título para PORTADA!Z5 y fila 8)
DOCUMENTO = (
    "Investigacion/Sistemas/listado_equipos.md",
    "P2437-HV-LIS-001",
    "LISTADO DE EQUIPOS Y MATERIALES (BOQ): SISTEMA DE VENTILACIÓN DEL LABORATORIO",
)

# ===================== ESTILOS CORPORATIVOS =====================
# Definidos en scripts/estilos_excel.py (importados arriba). Aquí solo queda
# la distribución de columnas específica de la BOQ.


def spans_tabla(n_cols):
    """Distribuye n_cols columnas markdown en el ancho total del encabezado
    corporativo (N_COLS columnas físicas).

    Para la BOQ del listado de equipos (7 columnas Markdown) se usa la
    distribución específica solicitada por el cliente:
      Ítem(1) + Función(2) + Especificación clave(3) + Material(2) +
      Cant.(1) + Candidatos comerciales(3) + Fuente(3) = 15 columnas.
    Para otras tablas se reparte lo más uniformemente posible."""
    if n_cols > N_COLS:
        raise ValueError(f"Tabla con {n_cols} columnas supera el layout A:{get_column_letter(N_COLS)}")
    if n_cols == 7:
        return [1, 2, 3, 2, 1, 3, 3]
    base, resto = divmod(N_COLS, n_cols)
    return [base + 1] * resto + [base] * (n_cols - resto)


# ===================== PARSEO MARKDOWN =====================
def limpiar_md(texto):
    """Quita sintaxis markdown: **negrita**, [texto](url) -> 'texto (url)', `código`."""
    texto = re.sub(r'\[([^\]]+)\]\(([^)]+)\)', r'\1 (\2)', texto)
    return texto.replace('**', '').replace('`', '').strip()


def es_separador_tabla(celdas):
    return all(re.fullmatch(r':?-{2,}:?', c.strip()) for c in celdas if c.strip()) \
        and any(c.strip() for c in celdas)


def parsear_linea_tabla(linea):
    return [limpiar_md(c) for c in linea.strip().strip('|').split('|')]


# ===================== CONSTRUCCIÓN DE LA HOJA =====================
class HojaLIS:
    def __init__(self, ws):
        self.ws = ws
        self.r = FILA_INICIO
        self._tras_blanco = False  # True si el elemento anterior ya dejó fila en blanco

    def _combinar(self, r1, r2, c1=1, c2=N_COLS):
        self.ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

    def titulo(self, texto, nivel):
        # Exactamente una fila en blanco antes de cada título (si el elemento
        # anterior no la dejó ya, p. ej. una tabla).
        if self.r > FILA_INICIO and not self._tras_blanco:
            self.r += 1
        self._tras_blanco = False
        cell = self.ws.cell(row=self.r, column=1, value=texto)
        cell.font = {1: title_font, 2: section_font}.get(nivel, subtitle_font)
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        self._combinar(self.r, self.r)
        self.r += 1

    def parrafo(self, texto, negrita=False):
        self._tras_blanco = False
        # Filas estimadas para el texto con wrap en el ancho total A:O
        ancho_total = sum(ANCHO_COLUMNAS)
        chars_linea = max(1.0, ancho_total / CHAR_POR_UNIDAD_ANCHO)
        n = min(max(1, math.ceil(len(texto) / chars_linea)), 15)
        cell = self.ws.cell(row=self.r, column=1, value=texto)
        cell.alignment = AL_PARRAFO
        if negrita:
            cell.font = caption_font
        else:
            cell.font = body_font
        if n > 1:
            self._combinar(self.r, self.r + n - 1)
        self.r += n

    def tabla(self, filas):
        spans = spans_tabla(len(filas[0]))
        hdr = self.r
        # Encabezado unificado
        col = 1
        for texto, n in zip(filas[0], spans):
            self.ws.cell(row=self.r, column=col, value=texto)
            if n > 1:
                self.ws.merge_cells(start_row=self.r, start_column=col,
                                    end_row=self.r, end_column=col + n - 1)
            # Estilo DESPUÉS de combinar (persiste en MergedCell)
            for cc in range(col, col + n):
                cell = self.ws.cell(row=self.r, column=cc)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = AL_HEADER
            col += n
        self.r += 1
        # Cuerpo
        for fila in filas[1:]:
            col = 1
            for texto, n in zip(fila, spans):
                cell = self.ws.cell(row=self.r, column=col, value=texto)
                cell.font = body_font
                cell.alignment = AL_TEXTO
                if n > 1:
                    self.ws.merge_cells(start_row=self.r, start_column=col,
                                        end_row=self.r, end_column=col + n - 1)
                col += n
            self.r += 1
        # Bordes thin en todo el rango A:O
        for row in self.ws.iter_rows(min_row=hdr, max_row=self.r - 1, min_col=1, max_col=N_COLS):
            for cell in row:
                cell.border = thin_border
        self.r += 1  # fila en blanco tras la tabla
        self._tras_blanco = True


def construir_lista(ws, md):
    hoja = HojaLIS(ws)
    lineas = md.split('\n')
    i = 0
    while i < len(lineas):
        ln = lineas[i].strip()
        i += 1
        if not ln or ln == '---':
            continue
        if ln.startswith('|'):
            bloque = [ln]
            while i < len(lineas) and lineas[i].strip().startswith('|'):
                bloque.append(lineas[i].strip())
                i += 1
            filas = [parsear_linea_tabla(l) for l in bloque]
            filas = [f for f in filas if not es_separador_tabla(f)]
            hoja.tabla(filas)
        elif ln.startswith('#'):
            nivel = len(ln) - len(ln.lstrip('#'))
            hoja.titulo(limpiar_md(ln.lstrip('#')), nivel)
        else:
            hoja.parrafo(limpiar_md(ln), negrita=ln.startswith('**Tabla'))


def insertar_encabezado(ws, ws_enc, imagenes, titulo_documento):
    """Replica el bloque corporativo de ENCABEZADO (filas 1-7) en `ws` y
    deja el título del documento en la fila 8."""
    # Filas 1-7: valores, estilos y alturas
    for rr in range(1, 8):
        ws.row_dimensions[rr].height = ws_enc.row_dimensions[rr].height
        for cc in range(1, 16):
            src = ws_enc.cell(row=rr, column=cc)
            dst = ws.cell(row=rr, column=cc)
            dst.value = src.value
            dst._style = copy(src._style)
    # Solo los merges de las filas 1-7 (el merge A8:O47 de la plantilla no se copia)
    for m in ws_enc.merged_cells.ranges:
        if m.min_row <= 7:
            ws.merge_cells(str(m))

    # Anchos proporcionales para todo el documento
    aplicar_ancho_columnas(ws)

    # Fuente corporativa en el encabezado copiado
    forzar_times_new_roman(ws, min_row=1, max_row=7)

    # Imágenes del encabezado
    for blob, anchor in imagenes:
        img = XLImage(PILImage.open(BytesIO(blob)))
        img.anchor = deepcopy(anchor)
        ws.add_image(img)

    # Fila 8: título del documento (sin el merge grande de la plantilla)
    ws.row_dimensions[8].height = ALTURA_MINIMA
    cell = ws.cell(row=8, column=1, value=titulo_documento)
    cell.font = subtitle_font
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.merge_cells(start_row=8, start_column=1, end_row=8, end_column=N_COLS)
    # Fuente corporativa también en la fila de título
    forzar_times_new_roman(ws, min_row=8, max_row=8)


def generar_libro(md_rel, codigo, titulo):
    md_path = ROOT / md_rel
    md = md_path.read_text(encoding='utf-8')

    wb = load_workbook(PLANTILLA)
    ws_enc = wb["ENCABEZADO"]
    imagenes = [(im._data(), deepcopy(im.anchor)) for im in ws_enc._images]

    # PORTADA: solo código y título; firmas/fecha/revisión quedan de la plantilla
    ws_portada = wb["PORTADA"]
    assert ws_portada["Z3"].value == "SISTEMA HVAC LABORATORIO BRINSA", ws_portada["Z3"].value
    assert ws_portada["N6"].value == "P2437", ws_portada["N6"].value
    ws_portada["BO1"] = codigo
    ws_portada["Z5"] = titulo

    ws = wb.create_sheet("LISTA")
    insertar_encabezado(ws, ws_enc, imagenes, titulo)
    construir_lista(ws, md)

    # Fuente corporativa en toda la portada
    forzar_times_new_roman(ws_portada)

    # Ajustar alturas de filas de contenido para evitar texto cortado
    ajustar_alturas_filas(ws, FILA_INICIO, ws.max_row)

    # Página A3 horizontal, ajustada a una página de ancho
    configurar_pagina_a3(ws)

    wb.remove(ws_enc)  # orden final: PORTADA, LISTA
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True

    SALIDA.mkdir(parents=True, exist_ok=True)
    out = SALIDA / f"{codigo} REV0.xlsx"
    wb.save(out)
    return out


def main():
    print("== GENERACIÓN DE LISTADO DE EQUIPOS (LIS) ==")
    md_rel, codigo, titulo = DOCUMENTO
    out = generar_libro(md_rel, codigo, titulo)
    print(f"  {codigo}  <-  {md_rel}")
    print(f"       ->  {out.relative_to(ROOT)}")
    print("OK: 1 libro LIS generado en build/lis/")


if __name__ == "__main__":
    main()
