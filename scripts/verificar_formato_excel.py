#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
verificar_formato_excel.py — Verificación read-back del formato corporativo
de los libros Excel generados (sin necesidad de Excel instalado).

Comprueba en cada hoja de contenido:
  1. Página A3 (paperSize = 8), orientación horizontal, fitToWidth = 1.
  2. Fuente Times New Roman 28 en el contenido (fila 9+).
  3. Encabezados de tabla con relleno verde claro DML (C6E0B4), sin azul.
  4. Alturas de fila >= ALTURA_MINIMA en filas con contenido.
  5. Sin dos filas vacías consecutivas (salvo zonas reservadas para imágenes).

Uso:  python scripts/verificar_formato_excel.py
"""

import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from estilos_excel import ALTURA_MINIMA, TAMANO, VERDE_CLARO  # noqa: E402

ROOT = Path(__file__).resolve().parent.parent

# (libro, hoja de contenido, fila de inicio del contenido)
LIBROS = [
    (ROOT / "memoriadecalculo.xlsx", "MEMORIA DE CÁLCULO", 9),
    (ROOT / "build" / "dts" / "P2437-HV-DTS-001 REV0.xlsx", "ESPECIFICACIÓN", 9),
    (ROOT / "build" / "dts" / "P2437-HV-DTS-002 REV0.xlsx", "ESPECIFICACIÓN", 9),
    (ROOT / "build" / "dts" / "P2437-HV-DTS-003 REV0.xlsx", "ESPECIFICACIÓN", 9),
    (ROOT / "build" / "lis" / "P2437-HV-LIS-001 REV0.xlsx", "LISTA", 9),
]

N_COLS = 15
errores = []


def filas_de_imagenes(ws):
    """Filas (1-based) cubiertas por imágenes ancladas en la hoja."""
    filas = set()
    for img in ws._images:
        try:
            f0 = img.anchor._from.row + 1
            alto_pt = img.height * 0.75 if img.height else 300
            n = int(alto_pt / ALTURA_MINIMA) + 2
            filas.update(range(f0, f0 + n))
        except AttributeError:
            pass
    return filas


def verificar(path, hoja, fila_inicio):
    nombre = path.name
    if not path.exists():
        errores.append(f"{nombre}: archivo no existe")
        return
    wb = load_workbook(path)
    ws = wb[hoja]

    # 1. Configuración de página
    ps = ws.page_setup
    if str(ps.paperSize) != "8":
        errores.append(f"{nombre}: paperSize={ps.paperSize} (esperado 8=A3)")
    if ps.orientation != "landscape":
        errores.append(f"{nombre}: orientation={ps.orientation} (esperado landscape)")
    if str(ps.fitToWidth) != "1":
        errores.append(f"{nombre}: fitToWidth={ps.fitToWidth} (esperado 1)")

    # 2-4. Fuente, rellenos y alturas en el contenido
    n_fuente_mal = 0
    n_relleno_azul = 0
    n_altura_baja = 0
    for row in ws.iter_rows(min_row=fila_inicio, max_row=ws.max_row, max_col=N_COLS):
        for cell in row:
            if cell.value is None:
                continue
            f = cell.font
            if f.name != "Times New Roman" or (f.size and f.size != TAMANO):
                n_fuente_mal += 1
            fill = cell.fill
            if fill and fill.fill_type == "solid":
                rgb = getattr(fill.start_color, "rgb", None)
                if isinstance(rgb, str) and rgb.endswith("1F4E78"):
                    n_relleno_azul += 1
        r = row[0].row
        if any(c.value is not None for c in row):
            h = ws.row_dimensions[r].height
            if h is not None and h < ALTURA_MINIMA:
                n_altura_baja += 1
    if n_fuente_mal:
        errores.append(f"{nombre}: {n_fuente_mal} celdas de contenido sin TNR {TAMANO}")
    if n_relleno_azul:
        errores.append(f"{nombre}: {n_relleno_azul} celdas con relleno azul 1F4E78")
    if n_altura_baja:
        errores.append(f"{nombre}: {n_altura_baja} filas con contenido y altura < {ALTURA_MINIMA}")

    # Encabezados de tabla: al menos una celda con verde claro DML
    verde_ok = any(
        cell.fill and cell.fill.fill_type == "solid"
        and isinstance(getattr(cell.fill.start_color, "rgb", None), str)
        and cell.fill.start_color.rgb.endswith(VERDE_CLARO)
        for row in ws.iter_rows(min_row=fila_inicio, max_row=ws.max_row, max_col=N_COLS)
        for cell in row
    )
    if not verde_ok:
        errores.append(f"{nombre}: no se encontró encabezado verde claro {VERDE_CLARO}")

    # 5. Filas vacías consecutivas (fuera de zonas de imágenes y de rangos
    # combinados con contenido, cuyas filas inferiores son MergedCell vacías)
    img_rows = filas_de_imagenes(ws)
    filas_con_contenido = set()
    for m in ws.merged_cells.ranges:
        if ws.cell(row=m.min_row, column=m.min_col).value is not None:
            filas_con_contenido.update(range(m.min_row, m.max_row + 1))
    vacias = 0
    for r in range(fila_inicio, ws.max_row + 1):
        if r in img_rows:
            vacias = 0
            continue
        if r in filas_con_contenido or any(
            ws.cell(row=r, column=c).value is not None for c in range(1, N_COLS + 1)
        ):
            vacias = 0
        else:
            vacias += 1
            if vacias >= 2:
                errores.append(f"{nombre}: filas vacías consecutivas terminan en fila {r}")
                break

    print(f"  {nombre}: verificado")


print("== VERIFICACIÓN DE FORMATO EXCEL ==")
for path, hoja, fila_inicio in LIBROS:
    verificar(path, hoja, fila_inicio)

if errores:
    print("\nERRORES:")
    for e in errores:
        print(f"  - {e}")
    sys.exit(1)
print("\nOK: formato corporativo verificado en los 5 libros.")
