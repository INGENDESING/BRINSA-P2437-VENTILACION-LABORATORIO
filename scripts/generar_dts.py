#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
generar_dts.py — Convierte las hojas de datos Markdown (Investigacion/Sistemas/
hojas_datos/*.md) a libros Excel con la plantilla corporativa
FormatosDocumentos/DTS.xlsx.

Cada libro resultante tiene EXACTAMENTE 2 hojas:
  - PORTADA: portada corporativa diligenciada (BO1 = código, Z5 = título).
  - ESPECIFICACIÓN: encabezado corporativo (filas 1-7, logos y fórmulas
    =PORTADA!...) y el contenido del .md apilado desde la fila 9, con todas
    las tablas distribuidas en el ancho total A:O (15 columnas físicas).

Salida: build/dts/<código> REV0.xlsx (intermedios que scripts/emitir.py copia
a Emisiones/3.0 HV-HOJAS DE DATOS/).

Uso:  python scripts/generar_dts.py
"""

import math
import re
import sys
from copy import copy, deepcopy
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment
from PIL import Image as PILImage

# Módulo único de formato corporativo DML (A3 horizontal, Times New Roman 28,
# encabezados verde claro, sin azul). Ver scripts/estilos_excel.py.
sys.path.insert(0, str(Path(__file__).resolve().parent))
from estilos_excel import (  # noqa: E402
    N_COLS, AL_TEXTO, AL_HEADER, AL_PARRAFO,
    header_fill, header_font, title_font, section_font, subtitle_font,
    caption_font, nota_font, body_font, thin_border,
    CHAR_POR_UNIDAD_ANCHO, ANCHO_COLUMNAS, ALTURA_LINEA,
    aplicar_ancho_columnas, spans_tabla, forzar_times_new_roman,
    ajustar_alturas_filas, configurar_pagina_a3,
)

ROOT = Path(__file__).resolve().parent.parent
PLANTILLA = ROOT / "FormatosDocumentos" / "DTS.xlsx"
SALIDA = ROOT / "build" / "dts"
FILA_INICIO = 9  # contenido tras encabezado corporativo (1-7) y fila 8 libre

# (archivo .md fuente, código del documento, título para PORTADA!Z5)
DOCUMENTOS = [
    ("Investigacion/Sistemas/hojas_datos/HD-VENT-001_ventilador.md",
     "P2437-HV-DTS-001",
     "HOJA DE DATOS Y ESPECIFICACIONES TÉCNICAS DEL VENTILADOR AXIAL TUBEAXIAL PRFV"),
    ("Investigacion/Sistemas/hojas_datos/HD-FILT-001_filtro_merv.md",
     "P2437-HV-DTS-002",
     "HOJA DE DATOS Y ESPECIFICACIONES TÉCNICAS DEL SISTEMA DE FILTRACIÓN MERV 13-14"),
    ("Investigacion/Sistemas/hojas_datos/HD-REJ-001_rejillas.md",
     "P2437-HV-DTS-003",
     "HOJA DE DATOS Y ESPECIFICACIONES TÉCNICAS DE LAS REJILLAS DE EXFILTRACIÓN"),
]

# ===================== ESTILOS CORPORATIVOS =====================
# Definidos en scripts/estilos_excel.py (importados arriba).


# ===================== PARSEO MARKDOWN =====================
def limpiar_md(texto):
    """Quita sintaxis markdown: **negrita**, [texto](url) -> 'texto (url)', `código`."""
    texto = re.sub(r'\[([^\]]+)\]\(([^)]+)\)', r'\1 (\2)', texto)
    return texto.replace('**', '').replace('`', '').strip()


def es_separador_tabla(celdas):
    return all(re.fullmatch(r':?-{2,}:?', c.strip()) for c in celdas if c.strip()) \
        and any(c.strip() for c in celdas)


def parsear_linea_tabla(linea):
    return [limpiar_md(c) for c in linea.strip().strip('|').split('|')]


# ===================== NOTAS NUMERADAS =====================
def dividir_notas_numeradas(texto):
    """Divide un párrafo que contiene varias notas numeradas N.N. en líneas separadas."""
    partes = re.split(r'(?=\b\d+\.\d+\.\s)', texto)
    return [p.strip() for p in partes if p.strip()]


def es_nota_numerada(texto):
    return bool(re.match(r'^\d+\.\d+\.\s', texto))


# ===================== CONSTRUCCIÓN DE LA HOJA =====================
class HojaDTS:
    def __init__(self, ws):
        self.ws = ws
        self.r = FILA_INICIO
        self._tras_blanco = False  # True si el elemento anterior ya dejó fila en blanco

    def _combinar(self, r1, r2, c2=N_COLS):
        self.ws.merge_cells(start_row=r1, start_column=1, end_row=r2, end_column=c2)

    def titulo(self, texto, nivel):
        # Exactamente una fila en blanco antes de cada título (si el elemento
        # anterior no la dejó ya, p. ej. una tabla).
        if self.r > FILA_INICIO and not self._tras_blanco:
            self.r += 1
        self._tras_blanco = False
        cell = self.ws.cell(row=self.r, column=1, value=texto)
        cell.font = {1: title_font, 2: section_font}.get(nivel, subtitle_font)
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        self._combinar(self.r, self.r)
        self.r += 1

    def _parrafo_simple(self, texto, negrita=False, fuente=None):
        # Filas estimadas para el texto con wrap en el ancho total A:O
        ancho_total = sum(ANCHO_COLUMNAS)
        chars_linea = max(1.0, ancho_total / CHAR_POR_UNIDAD_ANCHO)
        n = min(max(1, math.ceil(len(texto) / chars_linea)), 15)
        cell = self.ws.cell(row=self.r, column=1, value=texto)
        cell.alignment = AL_PARRAFO
        if fuente:
            cell.font = fuente
        elif negrita:
            cell.font = caption_font
        else:
            cell.font = body_font
        if n > 1:
            self._combinar(self.r, self.r + n - 1)
        self.r += n

    def parrafo(self, texto, negrita=False):
        self._tras_blanco = False
        if es_nota_numerada(texto):
            for nota in dividir_notas_numeradas(texto):
                self._parrafo_simple(nota, fuente=nota_font)
        else:
            self._parrafo_simple(texto, negrita=negrita)

    def tabla(self, filas):
        spans = spans_tabla(len(filas[0]))
        hdr = self.r
        # Encabezado unificado
        col = 1
        for texto, n in zip(filas[0], spans):
            self.ws.cell(row=self.r, column=col, value=texto)
            if n > 1:
                self.ws.merge_cells(start_row=self.r, start_column=col,
                                    end_row=self.r, end_column=col + n - 1)
            # Estilo DESPUÉS de combinar (persiste en MergedCell)
            for cc in range(col, col + n):
                cell = self.ws.cell(row=self.r, column=cc)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = AL_HEADER
            col += n
        self.r += 1
        # Cuerpo
        for fila in filas[1:]:
            col = 1
            for texto, n in zip(fila, spans):
                cell = self.ws.cell(row=self.r, column=col, value=texto)
                cell.font = body_font
                cell.alignment = AL_TEXTO
                if n > 1:
                    self.ws.merge_cells(start_row=self.r, start_column=col,
                                        end_row=self.r, end_column=col + n - 1)
                col += n
            self.r += 1
        # Bordes thin en todo el rango A:O
        for row in self.ws.iter_rows(min_row=hdr, max_row=self.r - 1, min_col=1, max_col=N_COLS):
            for cell in row:
                cell.border = thin_border
        self.r += 1  # fila en blanco tras la tabla
        self._tras_blanco = True


def construir_especificacion(ws, md):
    hoja = HojaDTS(ws)
    lineas = md.split('\n')
    i = 0
    while i < len(lineas):
        ln = lineas[i].strip()
        i += 1
        if not ln or ln == '---':
            continue
        if ln.startswith('|'):
            bloque = [ln]
            while i < len(lineas) and lineas[i].strip().startswith('|'):
                bloque.append(lineas[i].strip())
                i += 1
            filas = [parsear_linea_tabla(l) for l in bloque]
            filas = [f for f in filas if not es_separador_tabla(f)]
            hoja.tabla(filas)
        elif ln.startswith('#'):
            nivel = len(ln) - len(ln.lstrip('#'))
            hoja.titulo(limpiar_md(ln.lstrip('#')), nivel)
        else:
            hoja.parrafo(limpiar_md(ln), negrita=ln.startswith('**Tabla'))
    return hoja


def insertar_graficos_dts001(ws, fila):
    """Inserta la referencia gráfica del ventilador (solo DTS-001)."""
    img_dir = SALIDA / "img"
    curva = img_dir / "curva_ventilador_dts001.png"
    ref = img_dir / "ventilador_referencia_dts001.png"
    if not curva.exists() or not ref.exists():
        return fila
    # Título de sección
    cell = ws.cell(row=fila, column=1, value="Referencia gráfica")
    cell.font = section_font
    cell.alignment = Alignment(vertical='center', wrap_text=True)
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=N_COLS)
    fila += 2
    # Filas reservadas por imagen: 400 px ≈ 300 pt de alto / ALTURA_LINEA
    n_filas_img = math.ceil(400 * 0.75 / ALTURA_LINEA) + 1
    # Curva característica ilustrativa
    img = XLImage(str(curva))
    img.width = 600
    img.height = 400
    ws.add_image(img, f"A{fila}")
    fila += n_filas_img
    # Imagen de referencia del ventilador
    img = XLImage(str(ref))
    img.width = 600
    img.height = 400
    ws.add_image(img, f"A{fila}")
    fila += n_filas_img
    return fila


def insertar_encabezado(ws, ws_enc, imagenes):
    """Replica el bloque corporativo de ENCABEZADO (filas 1-7) en `ws`."""
    for rr in range(1, 8):
        ws.row_dimensions[rr].height = ws_enc.row_dimensions[rr].height
        for cc in range(1, 16):
            src = ws_enc.cell(row=rr, column=cc)
            dst = ws.cell(row=rr, column=cc)
            dst.value = src.value
            dst._style = copy(src._style)
    for m in ws_enc.merged_cells.ranges:
        if m.min_row <= 7:
            ws.merge_cells(str(m))
    # Anchos proporcionales para todo el documento
    aplicar_ancho_columnas(ws)
    # Fuente corporativa en el encabezado copiado
    forzar_times_new_roman(ws, min_row=1, max_row=7)
    for blob, anchor in imagenes:
        img = XLImage(PILImage.open(BytesIO(blob)))
        img.anchor = deepcopy(anchor)
        ws.add_image(img)


def generar_libro(md_rel, codigo, titulo):
    md_path = ROOT / md_rel
    md = md_path.read_text(encoding='utf-8')

    wb = load_workbook(PLANTILLA)
    ws_enc = wb["ENCABEZADO"]
    imagenes = [(im._data(), deepcopy(im.anchor)) for im in ws_enc._images]

    # PORTADA: solo código y título; firmas/fecha/revisión quedan de la plantilla
    ws_portada = wb["PORTADA"]
    assert ws_portada["Z3"].value == "SISTEMA HVAC LABORATORIO BRINSA", ws_portada["Z3"].value
    assert ws_portada["N6"].value == "P2437", ws_portada["N6"].value
    ws_portada["BO1"] = codigo
    ws_portada["Z5"] = titulo

    ws = wb.create_sheet("ESPECIFICACIÓN")
    hoja = construir_especificacion(ws, md)
    if codigo == "P2437-HV-DTS-001":
        insertar_graficos_dts001(ws, hoja.r)
    insertar_encabezado(ws, ws_enc, imagenes)

    # Fuente corporativa en toda la portada
    forzar_times_new_roman(ws_portada)

    # Ajustar alturas de filas de contenido para evitar texto cortado
    ajustar_alturas_filas(ws, FILA_INICIO, ws.max_row)

    # Página A3 horizontal, ajustada a una página de ancho
    configurar_pagina_a3(ws)

    wb.remove(ws_enc)  # orden final: PORTADA, ESPECIFICACIÓN
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True

    SALIDA.mkdir(parents=True, exist_ok=True)
    out = SALIDA / f"{codigo} REV0.xlsx"
    wb.save(out)
    return out


def main():
    print("== GENERACIÓN DE HOJAS DE DATOS (DTS) ==")
    for md_rel, codigo, titulo in DOCUMENTOS:
        out = generar_libro(md_rel, codigo, titulo)
        print(f"  {codigo}  <-  {md_rel}")
        print(f"       ->  {out.relative_to(ROOT)}")
    print("OK: 3 libros DTS generados en build/dts/")


if __name__ == "__main__":
    main()
