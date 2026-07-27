from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
from copy import copy, deepcopy
from io import BytesIO
import math
import os

# ===================== PLANTILLA CORPORATIVA =====================
# El libro nace de FormatosDocumentos/CAL.xlsx (PORTADA + ENCABEZADO).
# Resultado final: exactamente 2 hojas — PORTADA y MEMORIA DE CÁLCULO.
# En la hoja única, el encabezado corporativo ocupa las filas 1-7, la fila 8
# queda libre y el contenido se apila en secciones a partir de la fila 9.
RUTA_BASE = os.path.dirname(os.path.abspath(__file__))
RUTA_PLANTILLA = os.path.join(RUTA_BASE, "FormatosDocumentos", "CAL.xlsx")
FILA_INICIO = 9  # primera fila de contenido (tras encabezado 1-7 y fila 8 libre)

wb = load_workbook(RUTA_PLANTILLA)

# Capturar el bloque de encabezado (valores, estilos e imágenes) de la plantilla
ws_enc = wb["ENCABEZADO"]
ENC_IMAGENES = [(im._data(), deepcopy(im.anchor)) for im in ws_enc._images]

# Diligenciar PORTADA (BO2..BO6, Z1, Z3 y N6 ya vienen correctos en la plantilla)
ws_portada = wb["PORTADA"]
ws_portada["BO1"] = "P2437-HV-CAL-001"
ws_portada["Z5"] = "MEMORIA DE CÁLCULO DEL SISTEMA DE VENTILACIÓN DEL LABORATORIO"

# ===================== ESTILOS CORPORATIVOS =====================
# Fuente obligatoria para todo el documento: Times New Roman.
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(name="Times New Roman", color="FFFFFF", bold=True, size=11)
input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
input_font = Font(name="Times New Roman", bold=True, color="000000")
result_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
result_font = Font(name="Times New Roman", bold=True, color="006100")
formula_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
formula_font = Font(name="Times New Roman", color="000080", bold=True)
title_font = Font(name="Times New Roman", bold=True, size=14, color="1F4E78")
subtitle_font = Font(name="Times New Roman", bold=True, size=11, color="1F4E78")
body_font = Font(name="Times New Roman", size=11)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Alineaciones unificadas por rol
AL_TEXTO = Alignment(horizontal='left', vertical='top', wrap_text=True)
AL_CENTRO = Alignment(horizontal='center', vertical='center', wrap_text=True)
AL_HEADER = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Layout unificado: TODAS las tablas abarcan las 15 columnas del encabezado (A:O).
N_COLS = 15
FMT_RESULTADO = '0.0000'
CHAR_POR_UNIDAD_ANCHO = 2.3

# Anchos de columna proporcionales para el contenido A:O
ANCHO_COLUMNAS = [30.6, 12, 12, 6, 6, 6, 6, 6, 6, 8, 8, 8, 10, 10, 10]


def aplicar_ancho_columnas(ws):
    """Fija anchos proporcionales para las 15 columnas del documento."""
    for i, w in enumerate(ANCHO_COLUMNAS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def spans_tabla(n_cols):
    """Distribuye n_cols columnas lógicas en N_COLS columnas físicas (A:O).

    base = N_COLS // n_cols; el residuo se reparte en las primeras columnas.
    """
    if n_cols > N_COLS:
        raise ValueError(f"Tabla con {n_cols} columnas supera el layout A:{get_column_letter(N_COLS)}")
    base = N_COLS // n_cols
    extra = N_COLS % n_cols
    return [base + (1 if i < extra else 0) for i in range(n_cols)]


def col_inicio(n_cols, idx):
    """Columna física (1-based) donde empieza la columna lógica idx (1-based)."""
    spans = spans_tabla(n_cols)
    return 1 + sum(spans[:idx - 1])


def col_letra(n_cols, idx):
    """Letra de columna física donde empieza la columna lógica idx (1-based)."""
    return get_column_letter(col_inicio(n_cols, idx))


def forzar_times_new_roman(ws, min_row=1, max_row=None, min_col=1, max_col=N_COLS):
    """Sobrescribe el nombre de fuente conservando tamaño, negrita, cursiva, color y subrayado."""
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            f = cell.font
            cell.font = Font(
                name="Times New Roman",
                size=f.size,
                bold=f.bold,
                italic=f.italic,
                underline=f.underline,
                color=f.color,
            )


def apply_border(ws, start_row, start_col, end_row, end_col):
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.border = thin_border


def escribir_fila(ws, r, valores, alineaciones=None):
    """Escribe una fila de tabla. `valores`: lista de valores lógicos.
    `alineaciones`: dict índice lógico (1-based) -> Alignment."""
    n = len(valores)
    spans = spans_tabla(n)
    col = 1
    for i, valor in enumerate(valores):
        n_span = spans[i]
        cell = ws.cell(row=r, column=col, value=valor)
        cell.font = body_font
        cell.alignment = (alineaciones or {}).get(i + 1, AL_TEXTO)
        if n_span > 1:
            ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + n_span - 1)
        col += n_span


def encabezados_tabla(ws, r, headers):
    """Encabezado de tabla unificado (relleno, fuente, centrado).
    `headers`: lista de textos lógicos."""
    n = len(headers)
    spans = spans_tabla(n)
    col = 1
    for i, texto in enumerate(headers):
        n_span = spans[i]
        ws.cell(row=r, column=col, value=texto)
        if n_span > 1:
            ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + n_span - 1)
        # Estilo DESPUÉS de combinar: persiste en MergedCell.
        for cc in range(col, col + n_span):
            cell = ws.cell(row=r, column=cc)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = AL_HEADER
        col += n_span


def titulo_seccion(ws, r, texto, ultima_col=N_COLS):
    cell = ws.cell(row=r, column=1, value=texto)
    cell.font = title_font
    cell.alignment = AL_TEXTO
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ultima_col)


def subtitulo(ws, r, texto, ultima_col=N_COLS):
    cell = ws.cell(row=r, column=1, value=texto)
    cell.font = subtitle_font
    cell.alignment = AL_TEXTO
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ultima_col)


def texto_largo(ws, r_ini, r_fin, texto, ultima_col=N_COLS):
    cell = ws.cell(row=r_ini, column=1, value=texto)
    cell.font = body_font
    cell.alignment = AL_TEXTO
    ws.merge_cells(start_row=r_ini, start_column=1, end_row=r_fin, end_column=ultima_col)


def ajustar_alturas_filas(ws, primera, ultima):
    """Ajusta la altura de cada fila en función del texto más largo de la fila,
    considerando el ancho real (incluyendo celdas combinadas)."""
    from openpyxl.cell.cell import MergedCell
    for r in range(primera, ultima + 1):
        max_lineas = 1
        for c in range(1, N_COLS + 1):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            val = cell.value
            if not val:
                continue
            # Buscar rango combinado al que pertenezca la celda
            rng = None
            for m in ws.merged_cells.ranges:
                if cell.coordinate in m:
                    rng = m
                    break
            if rng:
                _, sc, _, ec = range_boundaries(str(rng))
                ancho = sum(ws.column_dimensions[get_column_letter(col)].width or 8.43
                            for col in range(sc, ec + 1))
                _, sr, _, _ = range_boundaries(str(rng))
                filas_combinadas = r - sr + 1
            else:
                ancho = ws.column_dimensions[get_column_letter(c)].width or 8.43
                filas_combinadas = 1
            chars_linea = max(1.0, ancho / CHAR_POR_UNIDAD_ANCHO)
            lineas_totales = math.ceil(len(str(val)) / chars_linea)
            lineas_por_fila = math.ceil(lineas_totales / filas_combinadas)
            max_lineas = max(max_lineas, lineas_por_fila)
        ws.row_dimensions[r].height = max(15, max_lineas * 15)


# ===================== HOJA ÚNICA: MEMORIA DE CÁLCULO =====================
ws = wb.create_sheet("MEMORIA DE CÁLCULO")
r = FILA_INICIO

# Helpers de columna para las tablas más usadas
C5 = lambda idx: col_letra(5, idx)   # sección 2 (entradas)
C6 = lambda idx: col_letra(6, idx)   # sección 3 (cálculos) y sección 6 (escenarios)

# ---------- SECCIÓN 1: DATOS GENERALES ----------
titulo_seccion(ws, r, "1. DATOS GENERALES — MEMORIA DE CÁLCULO, SISTEMA DE VENTILACIÓN")
r += 2  # fila en blanco
subtitulo(ws, r, "Información del proyecto")
r += 1
hdr = r
encabezados_tabla(ws, r, ["Concepto", "Valor", "Unidad", "Observación"])
r += 1
info = [
    ["Nombre del proyecto", "Laboratorio Brinsa", "-", "Ventilación por impulsión directa, descarga libre a atmósfera"],
    ["Ubicación", "C:\\Users\\ingen\\OneDrive\\Escritorio\\HVAC\\Calculos", "-", "Carpeta de cálculos del proyecto"],
    ["Volumen efectivo del laboratorio (V)", "320", "m³", "Medición / modelo 3D"],
    ["Renovaciones de aire (N)", "12", "ren/h", "Sustentado normativamente"],
    ["Objetivo", "Ventilación y filtración", "-", "12 ACH, MERV 13-14; sin presurización (REV1)"],
    ["Estrategia", "Ventilador axial mural Ø560 mm + cubierta intemperie con filtración + rejillas de exfiltración", "-", "Sin ductos de impulsión; montaje uniforme con la planta (REV2)"],
]
for row in info:
    escribir_fila(ws, r, row, alineaciones={3: AL_CENTRO})
    r += 1
apply_border(ws, hdr, 1, r - 1, N_COLS)
r += 1  # fila en blanco
subtitulo(ws, r, "Objetivo del documento")
r += 1
texto_largo(ws, r, r + 2,
    "Presentar de forma atómica, funcional y verificable el cálculo del caudal de aire, "
    "potencia del ventilador axial, área de impulsión del ventilador (sin ductos) y "
    "área/dimensiones de las rejillas de exfiltración para modelado CFD, con "
    "sustentación normativa.")
r += 4  # texto (3 filas) + fila en blanco

# ---------- SECCIÓN 2: ENTRADAS ----------
titulo_seccion(ws, r, "2. PARÁMETROS DE ENTRADA")
r += 2  # fila en blanco
hdr = r
encabezados_tabla(ws, r, ["Parámetro", "Símbolo", "Valor", "Unidad", "Referencia / Nota"])
r += 1
# Filas de entradas (columna Valor = C5(3))
e_V = r          # Volumen efectivo
e_N = r + 1      # Renovaciones de aire
e_dP = r + 2     # Presión total estimada del ventilador
e_eta = r + 3    # Eficiencia total
e_vvent = r + 4  # Velocidad en boca del ventilador
e_vsal = r + 5   # Velocidad de exfiltración
e_nrej = r + 6   # Número de rejillas
e_prop = r + 7   # Proporción ancho/alto
e_rho = r + 8    # Densidad del aire
e_Cd = r + 9     # Coeficiente de descarga
entradas = [
    ["Volumen efectivo del laboratorio", "V", 320, "m³", "Medición directa o del modelo"],
    ["Renovaciones de aire", "N", 12, "ren/h", "Ver sección 5. NORMATIVA"],
    ["Presión total estimada del ventilador", "ΔP", 165, "Pa", "Escenario MERV cargado (ver sección 6. ESCENARIOS DE FILTRACIÓN)"],
    ["Eficiencia total del ventilador", "η", 0.55, "-", "Axial típico (provisional; confirmar con catálogo)"],
    ["Velocidad en boca del ventilador", "v_vent", 8.0, "m/s", "Para CFD, rango 6–12 m/s"],
    ["Velocidad de exfiltración en rejillas", "v_sal", 3.0, "m/s", "Rango recomendado 2.5–4.0 m/s"],
    ["Número de rejillas de salida", "n_rej", 3, "unid.", "Distribución propuesta"],
    ["Proporción ancho/alto rejilla", "prop", 0.95, "-", "353 mm × 336 mm ≈ 0.95"],
    ["Densidad del aire", "rho", 0.88, "kg/m³", "Bases de diseño (Cajicá, 2 558 msnm, P_atm = 74.1 kPa, aire 20 °C)"],
    ["Coeficiente de descarga orificio", "Cd", 0.60, "-", "Orificio borde afilado, ASHRAE Handbook Fundamentals"],
]
for row in entradas:
    escribir_fila(ws, r, row, alineaciones={2: AL_CENTRO, 3: AL_CENTRO, 4: AL_CENTRO})
    # Destacar la columna Valor (lógica 3)
    for cc in range(col_inicio(5, 3), col_inicio(5, 3) + spans_tabla(5)[2]):
        c = ws.cell(row=r, column=cc)
        c.fill = input_fill
        c.font = input_font
    r += 1
apply_border(ws, hdr, 1, r - 1, N_COLS)
r += 1  # fila en blanco
subtitulo(ws, r, "Notas:")
r += 1
texto_largo(ws, r, r, "• Modifique los valores en amarillo; todos los cálculos se actualizan automáticamente.")
r += 1
texto_largo(ws, r, r, "• Las fórmulas están visibles en la barra de fórmulas de cada celda.")
r += 1
texto_largo(ws, r, r, "• Unidades: 1 m³/min = 35.3147 CFM; 1 kW = 1.341 HP.")
r += 2  # fila en blanco

# ---------- SECCIÓN 3: CÁLCULOS ----------
titulo_seccion(ws, r, "3. DESARROLLO DE CÁLCULOS")
r += 2  # fila en blanco
hdr = r
encabezados_tabla(ws, r, ["Paso", "Descripción", "Fórmula / Valor", "Resultado", "Unidad", "Referencia celda"])
r += 1
spans3 = spans_tabla(6)
res_col = col_inicio(6, 4)   # columna física del resultado
form_col = col_inicio(6, 3)  # columna física de la fórmula/valor
ref_col = col_inicio(6, 6)   # columna física de referencia celda
# Filas de resultados (columna Resultado = C6(4)) de cada paso
c_Q = r          # 1  Caudal m³/min
c_Qh = r + 1     # 2  Caudal m³/h
c_Qcfm = r + 2   # 3  Caudal CFM
c_Qs = r + 3     # 4  Caudal m³/s
c_P = r + 4      # 5  Potencia teórica kW
c_HP = r + 5     # 6  Potencia HP
c_Avent = r + 6  # 7  Área impulsión
c_D = r + 7      # 8  Diámetro equivalente
c_rad = r + 8    # 9  Radio
c_rmm = r + 9    # 10 Radio mm
c_Aexfil = r + 10 # 11 Área neta total rejillas
c_Arej = r + 11  # 12 Área neta por rejilla
c_h = r + 12     # 13 Altura rejilla
c_w = r + 13     # 14 Ancho rejilla
c_hmm = r + 14   # 15 Altura mm
c_wmm = r + 15   # 16 Ancho mm
c_dPrej = r + 16 # 17 Pérdida de presión en rejillas (descarga libre)
calculos = [
    ["1", "Caudal de aire volumétrico", "Q = V × N / 60", f"={C5(3)}{e_V}*{C5(3)}{e_N}/60", "m³/min"],
    ["2", "Caudal de aire en m³/h", "Q_h = Q × 60", f"={C6(4)}{c_Q}*60", "m³/h"],
    ["3", "Caudal de aire en CFM", "Q_cfm = Q × 35.3147", f"={C6(4)}{c_Q}*35.3147", "CFM"],
    ["4", "Caudal en m³/s", "Q_s = Q / 60", f"={C6(4)}{c_Q}/60", "m³/s"],
    ["5", "Potencia teórica del ventilador", "P = Q_s × ΔP / (η × 1000)", f"={C6(4)}{c_Qs}*{C5(3)}{e_dP}/({C5(3)}{e_eta}*1000)", "kW"],
    ["6", "Potencia en HP", "HP = P × 1.341", f"={C6(4)}{c_P}*1.341", "HP"],
    ["7", "Área de impulsión del ventilador", "A_vent = Q_s / v_vent", f"={C6(4)}{c_Qs}/{C5(3)}{e_vvent}", "m²"],
    ["8", "Diámetro equivalente del ventilador", "D = √(4×A_vent/π)", f"=SQRT(4*{C6(4)}{c_Avent}/PI())", "m"],
    ["9", "Radio del ventilador", "r = D / 2", f"={C6(4)}{c_D}/2", "m"],
    ["10", "Radio del ventilador en mm", "r_mm = r × 1000", f"={C6(4)}{c_rad}*1000", "mm"],
    ["11", "Área neta total de rejillas de salida", "A_exfil = Q_s / v_sal", f"={C6(4)}{c_Qs}/{C5(3)}{e_vsal}", "m²"],
    ["12", "Área neta por rejilla de salida", "A_rej = A_exfil / n_rej", f"={C6(4)}{c_Aexfil}/{C5(3)}{e_nrej}", "m²"],
    ["13", "Altura de rejilla (proporción dada)", "h = √(A_rej / prop)", f"=SQRT({C6(4)}{c_Arej}/{C5(3)}{e_prop})", "m"],
    ["14", "Ancho de rejilla", "w = A_rej / h", f"={C6(4)}{c_Arej}/{C6(4)}{c_h}", "m"],
    ["15", "Altura de rejilla en mm", "h_mm = h × 1000", f"={C6(4)}{c_h}*1000", "mm"],
    ["16", "Ancho de rejilla en mm", "w_mm = w × 1000", f"={C6(4)}{c_w}*1000", "mm"],
    ["17", "Pérdida de presión en rejillas (descarga libre)", "dP = (rho/2)×(Q_s/(Cd×A_exfil))²", f"={C5(3)}{e_rho}/2*({C6(4)}{c_Qs}/({C5(3)}{e_Cd}*{C6(4)}{c_Aexfil}))^2", "Pa"],
]
for row in calculos:
    paso_fila = r
    escribir_fila(ws, r, [row[0], row[1], row[2], row[3], row[4], f"={C6(4)}{paso_fila}"],
                  alineaciones={1: AL_CENTRO, 4: AL_CENTRO, 5: AL_CENTRO, 6: AL_CENTRO})
    # Fórmula / Valor (lógica 3): fondo gris y fuente de fórmula
    for cc in range(form_col, form_col + spans3[2]):
        c = ws.cell(row=r, column=cc)
        c.fill = formula_fill
        c.font = formula_font
    # Resultado (lógica 4): fondo verde, fuente resultado y formato numérico
    for cc in range(res_col, res_col + spans3[3]):
        c = ws.cell(row=r, column=cc)
        c.fill = result_fill
        c.font = result_font
        c.number_format = FMT_RESULTADO
    r += 1
apply_border(ws, hdr, 1, r - 1, N_COLS)
r += 1  # fila en blanco

# ---------- SECCIÓN 4: RESULTADOS ----------
titulo_seccion(ws, r, "4. RESUMEN DE RESULTADOS DE DISEÑO")
r += 2  # fila en blanco
hdr = r
encabezados_tabla(ws, r, ["Parámetro", "Valor", "Unidad", "Observación"])
r += 1
res_col4 = col_inicio(4, 2)
spans4 = spans_tabla(4)
resultados = [
    ["Caudal de diseño", f"={C6(4)}{c_Q}", "m³/min", "Caudal de impulsión requerido"],
    ["Caudal de diseño", f"={C6(4)}{c_Qh}", "m³/h", "Equivalente en metros cúbicos por hora"],
    ["Caudal de diseño", f"={C6(4)}{c_Qcfm}", "CFM", "Equivalente en pies cúbicos por minuto"],
    ["Potencia teórica del ventilador", f"={C6(4)}{c_P}", "kW", "Escenario MERV cargado a 165 Pa y η=0.55 (axial)"],
    ["Potencia teórica del ventilador", f"={C6(4)}{c_HP}", "HP", "Unidades imperiales"],
    ["Potencia instalada recomendada", f"=CEILING({C6(4)}{c_HP}*1.5,0.25)", "HP", "×1.5 margen de servicio, redondeo 0.25 HP (provisional; confirmar con catálogo)"],
    ["Área de impulsión del ventilador", f"={C6(4)}{c_Avent}", "m²", "Boca del ventilador"],
    ["Diámetro equivalente del ventilador", f"={C6(4)}{c_D}", "m", "Para referencia (BC del CFD, histórica)"],
    ["Radio del ventilador (CFD)", f"={C6(4)}{c_rmm}", "mm", "Círculo de inyección (BC del CFD, histórica)"],
    ["Diámetro del impulsor seleccionado", 0.56, "m", "Ø560 mm — uniformidad con montaje típico de planta (REV2)"],
    ["Área de boca real (Ø560 mm)", f"=PI()*(0.56/2)^2", "m²", "π × (D/2)²"],
    ["Velocidad real en boca (Ø560 mm)", f"={C6(4)}{c_Qs}/(PI()*(0.56/2)^2)", "m/s", "Q_s / A_boca real ≈ 4,33 m/s"],
    ["Velocidad de impulsión", f"={C5(3)}{e_vvent}", "m/s", "Condición de entrada CFD"],
    ["Área neta total de rejillas de salida", f"={C6(4)}{c_Aexfil}", "m²", "A v=3 m/s"],
    ["Número de rejillas de salida", f"={C5(3)}{e_nrej}", "unid.", "Distribución propuesta"],
    ["Área neta por rejilla de salida", f"={C6(4)}{c_Arej}", "m²", "Cada rejilla"],
    ["Altura de rejilla de salida", f"={C6(4)}{c_hmm}", "mm", "Dimensiones para CFD"],
    ["Ancho de rejilla de salida", f"={C6(4)}{c_wmm}", "mm", "Dimensiones para CFD"],
    ["Velocidad de exfiltración", f"={C5(3)}{e_vsal}", "m/s", "Condición de salida CFD (pressure outlet)"],
    ["Pérdida de presión en rejillas", f"={C6(4)}{c_dPrej}", "Pa", "Descarga libre a atmósfera, Cd=0.6"],
]
for row in resultados:
    escribir_fila(ws, r, row, alineaciones={2: AL_CENTRO, 3: AL_CENTRO})
    for cc in range(res_col4, res_col4 + spans4[1]):
        c = ws.cell(row=r, column=cc)
        c.fill = result_fill
        c.font = result_font
        c.number_format = FMT_RESULTADO
    r += 1
apply_border(ws, hdr, 1, r - 1, N_COLS)
r += 1  # fila en blanco

# ---------- SECCIÓN 5: NORMATIVA ----------
titulo_seccion(ws, r, "5. SUSTENTACIÓN NORMATIVA — 12 RENOVACIONES/HORA")
r += 2  # fila en blanco
hdr = r
encabezados_tabla(ws, r, ["Norma / Guía", "Año / Edición", "Recomendación ACH", "Observación"])
r += 1
normas = [
    ["ASHRAE 170 — Ventilation of Health Care Facilities", "2021", "6–15 ACH", "Laboratorios clínicos/procedimiento"],
    ["ASHRAE 62.1 — Ventilation for Acceptable IAQ", "2022", "Por ocupación y área", "Caudal mínimo de aire exterior"],
    ["OSHA 29 CFR 1910.1450 — Laboratory Standard", "2012", "Según PEL", "Ventilación para control de exposición"],
    ["NFPA 99 — Health Care Facilities Code", "2021", "Cumplimiento de sistemas", "Fiabilidad de sistemas de aire"],
    ["WHO — Laboratory Biosafety Manual (BSL-2)", "2004", "6–12 ACH", "12 ACH = límite superior BSL-2"],
    ["WHO — Laboratory Biosafety Manual (BSL-3)", "2004", "12–15 ACH", "12 ACH = umbral inferior BSL-3"],
    ["NIH — Design Requirements Manual", "2023", "10–12 ACH", "Punto de diseño para laboratorios biomédicos"],
    ["ASHRAE Handbook — Fundamentals", "2021", "—", "Coeficiente de descarga de orificios (Cd)"],
    ["RETIE (Colombia)", "vigente", "—", "Seguridad eléctrica del motor del ventilador"],
    ["NTC 2050 (Colombia)", "vigente", "—", "Circuitos y protecciones eléctricas"],
    ["Resolución 0312/2019 MinSalud (Colombia)", "2019", "—", "Habilitación de servicios de salud (si aplica)"],
]
for row in normas:
    escribir_fila(ws, r, row, alineaciones={2: AL_CENTRO, 3: AL_CENTRO})
    r += 1
apply_border(ws, hdr, 1, r - 1, N_COLS)
r += 1  # fila en blanco
subtitulo(ws, r, "Conclusión")
r += 1
texto_largo(ws, r, r + 2,
    "La selección de 12 renovaciones de aire por hora está sustentada por las guías "
    "internacionales de bioseguridad (WHO BSL-2/BSL-3, NIH DRM) y los estándares de "
    "ventilación de ASHRAE para instalaciones de salud. El marco regulatorio local "
    "(RETIE, NTC 2050, Resolución 0312/2019) aplica a la infraestructura y seguridad "
    "eléctrica del equipo; no existe norma colombiana que fije un valor de ACH distinto "
    "para este caso.")
r += 4  # texto (3 filas) + fila en blanco

# ---------- SECCIÓN 6: ESCENARIOS DE FILTRACIÓN ----------
titulo_seccion(ws, r, "6. ESCENARIOS DE FILTRACIÓN Y MOTOR RECOMENDADO")
r += 1
texto_largo(ws, r, r,
    f"ΔP_vent total = ΔP_filtro + ΔP_rejillas (11 Pa). Sin presurización (REV1: cambio "
    f"de alcance del cliente). El caudal Q_s se toma del paso 4 de la sección 3 "
    f"(celda {C6(4)}{c_Qs}). Sitio: Cajicá, Cundinamarca (2 558 msnm, ρ = 0.88 kg/m³).")
r += 2  # fila en blanco
hdr = r
headers6 = ["Escenario de filtración", "ΔP filtro (Pa)", "ΔP vent total (Pa)", "P teórica (kW)", "P teórica (HP)", "Motor recomendado (HP)"]
spans6 = spans_tabla(len(headers6))
col_idx6 = [1 + sum(spans6[:i]) for i in range(len(spans6))]
cols6 = [get_column_letter(ci) for ci in col_idx6]
encabezados_tabla(ws, r, headers6)
r += 1
escenarios = [
    ["MERV 13-14 limpio", 59],
    ["MERV 13-14 cargado (diseño)", 154],
    ["HEPA H13 limpio (referencia histórica — no aplica)", 250],
    ["HEPA H13 cargado (referencia histórica — no aplica)", 600],
]
for nombre, dp_filtro in escenarios:
    i = r
    escribir_fila(ws, i,
                  [nombre, dp_filtro, f"={cols6[1]}{i}+11",
                   f"={C6(4)}{c_Qs}*{cols6[2]}{i}/(0.55*1000)",
                   f"={cols6[3]}{i}*1.341",
                   f"=CEILING({cols6[4]}{i}*1.5,0.25)"],
                  alineaciones={2: AL_CENTRO, 3: AL_CENTRO, 4: AL_CENTRO, 5: AL_CENTRO, 6: AL_CENTRO})
    # Formato de las columnas de resultado (lógicas 2-6)
    fmts = {2: '0', 3: '0', 4: '0.000', 5: '0.000', 6: '0'}
    for log, fmt in fmts.items():
        sc = col_idx6[log - 1]
        for cc in range(sc, sc + spans6[log - 1]):
            c = ws.cell(row=i, column=cc)
            c.fill = result_fill
            c.font = result_font
            c.number_format = fmt
    r += 1
apply_border(ws, hdr, 1, r - 1, N_COLS)
r += 1  # fila en blanco
subtitulo(ws, r, "Nota:")
r += 1
texto_largo(ws, r, r + 2,
    "El motor provisional de 0.75 HP TEFC encapsulado anticorrosivo (en la corriente de "
    "aire, transmisión directa del ventilador mural Ø560 mm — uniformidad con el montaje "
    "típico de planta, REV2) corresponde al escenario MERV 13-14 "
    "cargado (punto de diseño), con margen de servicio 1.5 sobre la potencia teórica; la "
    "potencia definitiva se fija con la curva del ventilador axial seleccionado. Los "
    "escenarios HEPA son solo referencia histórica: el laboratorio de análisis industrial "
    "NO requiere HEPA. Para selección de ventilador en catálogo (ρ = 1.2 kg/m³), usar el "
    "punto equivalente 3 840 m³/h @ 225 Pa (diseño) o 95 Pa (limpio). La potencia se "
    "recalcula automáticamente si cambia el caudal en la sección 3. DESARROLLO DE CÁLCULOS.")
r += 4  # texto (3 filas) + fila en blanco

# ---------- SECCIÓN 7: VISTA RÁPIDA ----------
titulo_seccion(ws, r, "7. VISTA RÁPIDA — RESULTADOS (FÓRMULAS VIVAS)")
r += 2  # fila en blanco
hdr = r
encabezados_tabla(ws, r, ["Parámetro", "Valor", "Unidad", "Observación"])
r += 1
res_col7 = col_inicio(4, 2)
spans7 = spans_tabla(4)
vista_rapida = [
    ["Caudal de diseño", f"={C6(4)}{c_Q}", "m³/min", "Caudal de impulsión requerido"],
    ["Caudal de diseño", f"={C6(4)}{c_Qh}", "m³/h", "Equivalente en metros cúbicos por hora"],
    ["Caudal de diseño", f"={C6(4)}{c_Qcfm}", "CFM", "Equivalente en pies cúbicos por minuto"],
    ["Potencia teórica del ventilador", f"={C6(4)}{c_P}", "kW", "Escenario MERV cargado a 165 Pa (axial)"],
    ["Potencia teórica del ventilador", f"={C6(4)}{c_HP}", "HP", "Unidades imperiales"],
    ["Potencia instalada recomendada", f"=CEILING({C6(4)}{c_HP}*1.5,0.25)", "HP", "×1.5 margen de servicio (provisional)"],
    ["Área de impulsión del ventilador", f"={C6(4)}{c_Avent}", "m²", "Boca del ventilador"],
    ["Diámetro equivalente del ventilador", f"={C6(4)}{c_D}", "m", "Para referencia (BC del CFD, histórica)"],
    ["Radio del ventilador (CFD)", f"={C6(4)}{c_rmm}", "mm", "Círculo de inyección (BC del CFD, histórica)"],
    ["Diámetro del impulsor seleccionado", 0.56, "m", "Ø560 mm — uniformidad con montaje típico de planta (REV2)"],
    ["Área de boca real (Ø560 mm)", f"=PI()*(0.56/2)^2", "m²", "π × (D/2)²"],
    ["Velocidad real en boca (Ø560 mm)", f"={C6(4)}{c_Qs}/(PI()*(0.56/2)^2)", "m/s", "Q_s / A_boca real ≈ 4,33 m/s"],
    ["Velocidad de impulsión", f"={C5(3)}{e_vvent}", "m/s", "Condición de entrada CFD"],
    ["Área neta total de rejillas de salida", f"={C6(4)}{c_Aexfil}", "m²", "A v=3 m/s"],
    ["Número de rejillas de salida", f"={C5(3)}{e_nrej}", "unid.", "Distribución propuesta"],
    ["Área neta por rejilla de salida", f"={C6(4)}{c_Arej}", "m²", "Cada rejilla"],
    ["Altura de rejilla de salida", f"={C6(4)}{c_hmm}", "mm", "Dimensiones para CFD"],
    ["Ancho de rejilla de salida", f"={C6(4)}{c_wmm}", "mm", "Dimensiones para CFD"],
    ["Velocidad de exfiltración", f"={C5(3)}{e_vsal}", "m/s", "Condición de salida CFD"],
    ["Pérdida de presión en rejillas", f"={C6(4)}{c_dPrej}", "Pa", "Descarga libre a atmósfera, Cd=0.6"],
]
for row in vista_rapida:
    escribir_fila(ws, r, row, alineaciones={2: AL_CENTRO, 3: AL_CENTRO})
    for cc in range(res_col7, res_col7 + spans7[1]):
        c = ws.cell(row=r, column=cc)
        c.fill = result_fill
        c.font = result_font
        c.number_format = FMT_RESULTADO
    r += 1
apply_border(ws, hdr, 1, r - 1, N_COLS)
r += 1  # fila en blanco
subtitulo(ws, r, "Nota:")
r += 1
texto_largo(ws, r, r + 2,
    "Sección con fórmulas vivas que referencian las secciones 3. DESARROLLO DE CÁLCULOS y "
    "2. PARÁMETROS DE ENTRADA de esta misma hoja. Si modifica las entradas, esta vista y la "
    "sección 4. RESULTADOS se actualizan automáticamente.")

# ===================== ENCABEZADO CORPORATIVO (filas 1-7) =====================
for rr in range(1, 8):
    ws.row_dimensions[rr].height = ws_enc.row_dimensions[rr].height
    for cc in range(1, 16):
        src = ws_enc.cell(row=rr, column=cc)
        dst = ws.cell(row=rr, column=cc)
        dst.value = src.value
        dst._style = copy(src._style)
for m in ws_enc.merged_cells.ranges:
    if m.min_row <= 7:
        ws.merge_cells(str(m))

# Anchos proporcionales para todo el documento (sobrescriben el ancho por defecto)
aplicar_ancho_columnas(ws)

# Forzar Times New Roman en el encabezado corporativo copiado
forzar_times_new_roman(ws, min_row=1, max_row=7)

# Imágenes del encabezado
for blob, anchor in ENC_IMAGENES:
    img = XLImage(PILImage.open(BytesIO(blob)))
    img.anchor = deepcopy(anchor)
    ws.add_image(img)

# Ajustar alturas de fila del contenido para evitar texto cortado
ajustar_alturas_filas(ws, FILA_INICIO, r + 2)

# Forzar Times New Roman en toda la portada
forzar_times_new_roman(ws_portada)

# Eliminar la hoja de ejemplo; orden final: PORTADA, MEMORIA DE CÁLCULO
wb.remove(ws_enc)
wb.move_sheet("MEMORIA DE CÁLCULO", offset=-(wb.sheetnames.index("MEMORIA DE CÁLCULO") - 1))

# ===================== CONFIGURACIÓN DE CÁLCULO =====================
wb.calculation.calcMode = "auto"
wb.calculation.fullCalcOnLoad = True

# ===================== GUARDAR =====================
output_path = os.path.join(RUTA_BASE, "memoriadecalculo.xlsx")
wb.save(output_path)
print(f"Archivo guardado en: {output_path}")
print(f"Hojas: {wb.sheetnames}")
